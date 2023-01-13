import json
from pathlib import Path
import random
import time
from colorama import Fore, init
import requests
import openpyxl
import js2py
import os
from dotenv import dotenv_values
import cups
import pytesseract
import cv2


init()
config = dotenv_values('.env')
start_message = f"""
{Fore.MAGENTA}
Что вы хотите сделать?
[1] - Добавить задания в отправку
[2] - Расставить посылки по ячейкам
[3] - Получить смс от отправления
[4] - Выдать посылку если забыли сказать SMS-код
[5] - Где посылка?
[6] - Печать этикетки
"""
parse_code = js2py.eval_js("""
function parse_code(code) {
    const route = +('' + code).substring(0, 1);
    const payload = '' + parseInt(code.substring(1));
    const id = +payload.substring(0, payload.length - 1);
    return id;
}
""")
BASE_DIR = Path(__file__).resolve().parent
SHIPMENT_STATUS_VALUES = {
  0: "Импортирован",
  10: "Принят от отправителя",
  20: "Готов к отгрузке",
  30: "Зарегистрирована накладная",
  40: "Выдано курьеру",
  45: "В пути из ПВЗ",
  50: "Прибыл в РЦ",
  60: "Передано в транзит \"Город отправления\" - \"Город получения\"",
  70: "Поступил в РЦ город получения",
  75: "В пути на ПВЗ",
  80: "Готов к выдаче",
  90: "Выдано",
  100: "Отменен",
  110: "Возврат",
  140: "Не прибыл на ПВЗ (недостача)",
  150: "Излишек",
  151: "Засыл",
  91: "Выдано частично"
}

labelPrinter = "_Label_Printer_2"
invoicePrinter = "HP_LaserJet_Pro_MFP_M127fn"


def build_headers(session: requests.Session):
    return {
        'Authorization': f"Bearer {session.cookies.get('Bearer')}",
        'X-XSRF-TOKEN': session.cookies.get("XSRF-TOKEN"),
        'Accept': 'application/json, text/plain, */*',
        'Origin': 'https://pvz.exmail24.ru',
        'Referer': 'https://pvz.exmail24.ru/',
        'Content-Type': 'application/json',
        'Accept-Language': 'ru',
        'Host': 'pvz.exmail24.ru',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive'
    }


def login(login_data):
    session = requests.Session()
    response_bearer = session.post("https://pvz.exmail24.ru/api/sanctum/token", data=login_data)
    session.cookies.set("Bearer", response_bearer.json()['token'])
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Origin': 'https://pvz.exmail24.ru',
        'Authorization': f"Bearer {response_bearer.json()['token']}",
        'Referer': 'https://pvz.exmail24.ru/',
        'Content-Type': 'application/json',
        'Accept-Language': 'ru',
        'Host': 'pvz.exmail24.ru',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive'
    }
    session.post("https://pvz.exmail24.ru/api/sanctum/user", headers=headers)
    return session


def get_warehouse(session: requests.Session, warehouse_id):
    return session.get(f"https://pvz.exmail24.ru/api/warehouse/{warehouse_id}", headers=build_headers(session), cookies=session.cookies)


def get_freight(session: requests.Session, freight_id):
    return session.get(f"https://pvz.exmail24.ru/api/freights/{freight_id}", headers=build_headers(session), cookies=session.cookies)


def get_shipment(session: requests.Session, shipment_id):
    return session.get(f"https://pvz.exmail24.ru/api/shipments/{shipment_id}", headers=build_headers(session), cookies=session.cookies)


def send_shipment_sms(session: requests.Session, shipment_id):
    return session.get(f"https://pvz.exmail24.ru/api/shipments/{shipment_id}/sms", headers=build_headers(session), cookies=session.cookies)


def issued_shipment(session: requests.Session, shipment_id, code):
    return session.put(f"https://pvz.exmail24.ru/api/shipments/{shipment_id}/issued", headers=build_headers(session), cookies=session.cookies, data=json.dumps({"sms": str(code)}))


def place_shipment(session: requests.Session, shipment_id, shipment_data):
    return session.put(f"https://pvz.exmail24.ru/api/shipments/{shipment_id}/placed", headers=build_headers(session), cookies=session.cookies, data=json.dumps(shipment_data))


def dump_shipment(session: requests.Session, shipments, sending):
    return session.put("https://pvz.exmail24.ru/api/sendings/" + str(sending), headers=build_headers(session), cookies=session.cookies, data=json.dumps({"shipments": shipments}))


def get_ceils(session: requests.Session):
    return session.get("https://pvz.exmail24.ru/api/ceils?", headers=build_headers(session), cookies=session.cookies)


def get_sticker(session: requests.Session, shipment_id):
    return session.get(f"https://pvz.exmail24.ru/api/etic-pdf?shipment_id={shipment_id}", headers=build_headers(session), cookies=session.cookies)


def get_acceptance(session: requests.Session, shipment_id):
    return session.get(f"https://pvz.exmail24.ru/api/receipt_avito_acceptance?shipment_id={shipment_id}", headers=build_headers(session), cookies=session.cookies)


def get_shipment_russian(shipment_id):
    return requests.get(f"https://sa.exmail24.ru/api/shipments/get-id/{shipment_id}")


def sort_accept(login_data, freight_id):
    wb = openpyxl.load_workbook(F'{BASE_DIR}/input/accept.xlsx')
    sheet = wb.active
    shipments_to_add = []
    sents = []
    for row in range(1, sheet.max_row, 2):
        if sheet.cell(row=row, column=1).value is not None:
            shipments_to_add.append({
                "ceil_id": parse_code(str(round(sheet.cell(row=row+1, column=1).value))),
                "shipment_id": parse_code(str(round(sheet.cell(row=row, column=1).value)))
            })
    session = login(login_data)
    for i, shipment_data in enumerate(shipments_to_add):
        place_shipment(session, shipment_data['shipment_id'], {"ceil_id": shipment_data['ceil_id'], "freight_id": freight_id})
        shipment = get_shipment(session, shipment_data['shipment_id'])
        shipment_json = shipment.json()
        while shipment.status_code == 429:
            # print(Fore.LIGHTRED_EX + "[WARNING] Много запросов статус: {}".format(shipment.status_code))
            time.sleep(5)
            shipment = get_shipment(session, shipment_data['shipment_id'])
            shipment_json = shipment.json()
            if shipment.status_code != 429:
                break
        if shipment_json.get('point_dst', {}).get('id', False) != 275:
            sents.append(shipment_data)
            print(Fore.LIGHTRED_EX + f"#{i+1} Засыл {shipment_json['number']}({shipment_json['id']}) - полка {shipment_json['ceil']['name']}")
        else:
            shipment = get_shipment(session, shipment_data['shipment_id'])
            shipment_json = shipment.json()
            if shipment_json.get('status', '') == "150":
                place_shipment(session, shipment_data['shipment_id'], {"ceil_id": shipment_data['ceil_id'], "freight_id": freight_id})
                shipment_json = get_shipment(session, shipment_data['shipment_id']).json()
                print(Fore.LIGHTCYAN_EX + f"#{i+1} {shipment_json['number']}({shipment_json['id']}) {shipment_json['ceil']['name']} успешно размещена X2")
            else:
                print(Fore.LIGHTCYAN_EX + f"#{i+1} [{shipment.status_code}] {shipment_json.get('number', '')}({shipment_json.get('id', '')}) {shipment_json.get('ceil', {}).get('name', 'Ошибка')} успешно размещена")


def sort_send(session, send_id):
    wb = openpyxl.load_workbook(F'{BASE_DIR}/input/add.xlsx')
    sheet = wb.active
    shipments_to_add = []
    for row in range(1, sheet.max_row+1):
        if sheet.cell(row=row, column=1).value is not None:
            shipments_to_add.append(parse_code(str(sheet.cell(row=row, column=1).value)))
    for i in range(0, len(shipments_to_add), 8):
        response = dump_shipment(session, shipments_to_add[i:i+8], send_id)
        if response.status_code == 200:
            print(Fore.LIGHTGREEN_EX + f"[INFO] Добавлено {len(shipments_to_add[0:i+8])} из {len(shipments_to_add)}")
        else:
            print(Fore.RED + f"[ERROR] {response.text}")


def get_sticker_file(session, shipment_id):
    sticker = get_sticker(session, shipment_id)
    with open(f"{BASE_DIR}/temp/sticker.pdf", 'wb') as file:
        file.write(sticker.content)


def get_invoice_file(session, shipment_id):
    invoice = get_acceptance(session, shipment_id)
    with open(f"{BASE_DIR}/temp/invoice.xlsx", 'wb') as file:
        file.write(invoice.content)


def print_sticker_file():
    conn = cups.Connection()
    labelPrinterOptions = {
        "PrintSpeed": "60",
        "PageSize": "w295h417",
        "Darkness": "5",
        "Vertical": "5"
    }
    conn.printFile(labelPrinter, f"{BASE_DIR}/temp/sticker.pdf", "Label", labelPrinterOptions)
    os.remove(f"{BASE_DIR}/temp/sticker.pdf")


def print_invoice_file():
    pass
    # TODO: печать накладной
    # conn.printFile(invoicePrinter, f"{BASE_DIR}/temp/invoice.pdf", "Invoice", {})
    # os.remove(f"{BASE_DIR}/temp/invoice.xlsx")


def check_shipments(shipments):
    if config.get("EXMAIL_PASSWORD", False) is False or config.get("EXMAIL_PASSWORD", False) == '' or config.get("EXMAIL_LOGIN", False) is False or config.get("EXMAIL_LOGIN", False) == '':
        print(Fore.RED + "[ERROR] Не указанны переменный в файле .env, укажите логин(почту) и пароль от exmail")
        exit()
    login_data = {"password": config['EXMAIL_PASSWORD'], "email_adress": config['EXMAIL_LOGIN'], "remember": True}
    session = login(login_data)
    for i, shipment_code in enumerate(shipments):
        if len(str(shipment_code)) == 13:
            shipment_code = parse_code(str(shipment_code))
        shipment = get_shipment(session=session, shipment_id=shipment_code)
        while shipment.status_code == 429:
            time.sleep(2)
            shipment = get_shipment(session=session, shipment_id=shipment_code)
        try:
            if int(shipment.json().get("status")) == 90:
                print(Fore.LIGHTGREEN_EX + "[SUCCESS] #{}. {}({}) статуc {}, пвз {}".format(i+1, shipment.json().get("number", ""), shipment.json().get("id", ""), SHIPMENT_STATUS_VALUES[int(shipment.json().get("status"))], shipment.json().get("dts_point_id", "")))
            elif int(shipment.json().get("status")) == 100:
                print(Fore.LIGHTBLUE_EX + "[SUCCESS] #{}. {}({}) статуc {}, пвз {}".format(i+1, shipment.json().get("number", ""), shipment.json().get("id", ""), SHIPMENT_STATUS_VALUES[int(shipment.json().get("status"))], shipment.json().get("dts_point_id", "")))
            else:
                print(Fore.RED + "[WARNING] #{}. {}({}) статуc {}, пвз {}".format(i+1, shipment.json().get("number", ""), shipment.json().get("id", ""), SHIPMENT_STATUS_VALUES[int(shipment.json().get("status"))], shipment.json().get("dts_point_id", "")))
        except TypeError:
            print(Fore.RED + "[ERROR] #{}. {}".format(i+1, shipment_code))


def decode_shipment_code(shipment_id):
    try:
        shipment_id = str(shipment_id).strip().replace("\n", '').replace("-", '')
        if len(str(shipment_id)) == 13:
            return parse_code(str(shipment_id))
        shipment_response = get_shipment_russian(shipment_id)
        if shipment_response.json() == {}:
            return "Ошибка {}".format(shipment_id)
        return int(shipment_response.text)
    except Exception:
        return "Ошибка {}".format(shipment_id)


def decode_shipments_from_photo(path):
    image = cv2.imread(path)
    string = pytesseract.image_to_string(image,  config='digits')
    shipments = []
    for shipment in string.split("\n"):
        if len(str(shipment)) != 0:
            shipments.append(decode_shipment_code(shipment))
    return shipments


def main():
    try:
        if config.get("EXMAIL_PASSWORD", False) is False or config.get("EXMAIL_PASSWORD", False) == '' or config.get("EXMAIL_LOGIN", False) is False or config.get("EXMAIL_LOGIN", False) == '':
            print(Fore.RED + "[ERROR] Не указанны переменный в файле .env, укажите логин(почту) и пароль от exmail")
            exit()
        login_data = {"password": config['EXMAIL_PASSWORD'], "email_adress": config['EXMAIL_LOGIN'], "remember": True}
        while True:
            try:
                task_type = int(input(start_message))
                session = login(login_data)
                if task_type == 1:
                    print(Fore.LIGHTWHITE_EX + "[INFO] Вы выбрали [1] - Добавить задания в отправку\n\nУбедитесь что файл add.xlsx лежит в папке с скриптом, в таблице в первом столбике подряд должны идти шк отправлений, проще всего отсканировать сканнером в таблицу.\n\n" + Fore.WHITE + "(для выхода назад напишите 'Назад')\n\n")
                    while True:
                        try:
                            warehouse_id = input(Fore.LIGHTWHITE_EX + "[INPUT DATA] Введите номер отправки: ")
                            if str(warehouse_id).lower() == 'назад':
                                break
                            print(Fore.LIGHTBLUE_EX + "[INFO] Обрабатываю Excel файл с ШК посылок на отправку")
                            sort_send(session, warehouse_id)
                            print(Fore.GREEN + "[SUCCESS] Обработка завершена")
                            break
                        except ValueError:
                            print(Fore.LIGHTRED_EX + '[ERROR] Неверный номер отправки, убедитесь что вы вводите номер отправки который указан в ссылке, попробуйте еще раз' + Fore.RESET)
                            continue
                elif task_type == 2:
                    print(Fore.LIGHTWHITE_EX + "[INFO] Вы выбрали [2] - Расставить посылки по ячейкам\n\nУбедитесь что файл accept.xlsx лежит в папке с скриптом, в таблице в первом столбике подряд должны чередоваться шк отправлений и шк половок(пример: 1000010143130, 4000000001379), проще всего отсканировать сканнером в таблицу.\n\n" + Fore.WHITE + "(для выхода назад напишите 'Назад')\n\n")
                    while True:
                        try:
                            freight_id = input(Fore.LIGHTWHITE_EX + "[INPUT DATA] Введите номер перевозки: ")
                            if str(freight_id).lower() == 'назад':
                                break
                            freight_id = int(freight_id)
                            response = get_freight(session, freight_id)
                            if response.status_code == 404:
                                raise ValueError
                            print(Fore.LIGHTBLUE_EX + "[INFO] Размещаю посылки по полочкам")
                            sort_accept(login_data, freight_id)
                            print(Fore.GREEN + "[SUCCESS] Обработка завершена")
                            break
                        except ValueError:
                            print(Fore.LIGHTRED_EX + '[ERROR] Неверный номер перевозки, убедитесь что вы вводите номер перевозки который указан в ссылке, попробуйте еще раз' + Fore.RESET)
                            continue
                elif task_type == 3:
                    print(Fore.LIGHTWHITE_EX + "[INFO] Вы выбрали [3] - Получить смс от отправления\n\nУбедитесь что вы выслали смс и у вас открыто окно с вводом смс. Будьте внимательны, если выслать код еще раз, то он измениться!\n\n" + Fore.WHITE + "(для выхода назад напишите 'Назад')\n\n")
                    while True:
                        try:
                            shipment_id = input(Fore.LIGHTWHITE_EX + "[INPUT] Введите шк отправления: ")
                            if str(shipment_id).lower() == 'назад':
                                break
                            shipment_id = int(shipment_id)
                            if len(str(shipment_id)) == 13:
                                shipment_id = parse_code(str(shipment_id))
                            shipment = get_shipment(session=session, shipment_id=shipment_id)
                            if shipment.status_code == 404:
                                raise ValueError
                            sms = shipment.json()["sms"]
                            if sms is None:
                                print(Fore.LIGHTYELLOW_EX + '[WARNING] Вы не выслали SMS-код, попробуйте еще раз' + Fore.RESET)
                            else:
                                print(Fore.LIGHTGREEN_EX + f'[SUCCESS] SMS-код сообщение от отправления: {shipment.json()["sms"]}')
                                break
                        except ValueError:
                            print(Fore.LIGHTRED_EX + '[ERROR] Неверное шк отправления, попробуйте еще раз' + Fore.RESET)
                            continue
                elif task_type == 4:
                    print(Fore.LIGHTWHITE_EX + "[INFO] Вы выбрали [4] - Выдать посылку если забыли сказать SMS-код\n\nБудьте внимательны, действие необратимо!\n\n" + Fore.WHITE + "(для выхода назад напишите 'Назад')\n\n")
                    while True:
                        try:
                            shipment_id = input(Fore.LIGHTWHITE_EX + "[INPUT] Введите шк отправления: ")
                            if str(shipment_id).lower() == 'назад':
                                break
                            shipment_id = int(shipment_id)
                            if len(str(shipment_id)) == 13:
                                shipment_id = parse_code(str(shipment_id))
                            shipment = get_shipment(session=session, shipment_id=shipment_id)
                            if shipment.status_code == 404:
                                raise ValueError
                            sms = shipment.json()["sms"]
                            if sms is None:
                                send_shipment_sms(session, shipment_id)
                                shipment = get_shipment(session=session, shipment_id=shipment_id)
                                sms = shipment.json()["sms"]
                            verificate_code = random.randint(0, 9999)
                            ready = input(Fore.RED + f"[WARNING!] Вы уверены что хотите выдать отправление {shipment.json()['number']}({shipment.json()['id']}). Для подтверждения введите {verificate_code}\n" + Fore.WHITE + "(для выхода назад напишите 'Назад')\n\n")
                            if str(ready).lower() == 'назад':
                                break
                            elif str(ready) == str(verificate_code):
                                issued_shipment(session, shipment_id, sms)
                                print(Fore.LIGHTGREEN_EX + f"[SUCCESS] Отправление {shipment.json()['number']}({shipment.json()['id']}) успешно выдано!")
                                break
                            else:
                                raise KeyboardInterrupt
                        except ValueError:
                            print(Fore.LIGHTRED_EX + '[ERROR] Неверное шк отправления, попробуйте еще раз' + Fore.RESET)
                            continue
                elif task_type == 5:
                    print(Fore.LIGHTWHITE_EX + "[INFO] Вы выбрали [5] - Где посылка?\n\n" + Fore.WHITE + "(для выхода назад напишите 'Назад')\n\n")
                    while True:
                        try:
                            shipment_id = input(Fore.LIGHTWHITE_EX + "[INPUT] Введите шк отправления: ")
                            if str(shipment_id).lower() == 'назад':
                                break
                            shipment_id = int(shipment_id)
                            if len(str(shipment_id)) == 13:
                                shipment_id = parse_code(str(shipment_id))
                            shipment = get_shipment(session=session, shipment_id=shipment_id)
                            if shipment.status_code == 404:
                                raise ValueError
                            ceils = get_ceils(session).json()
                            if shipment.json()['ceil_id'] is not None:
                                for ceil in ceils['ceils']:
                                    if ceil['id'] == shipment.json()['ceil_id']:
                                        print(Fore.LIGHTGREEN_EX + f'[SUCCESS] Отправление: {shipment.json()["number"]} - полка {ceil["name"]}')
                                        break
                            else:
                                print(Fore.LIGHTRED_EX + f'[ERROR] У отправления {shipment.json()["number"]} нет полки, статус {SHIPMENT_STATUS_VALUES[int(shipment.json()["status"])]}' + Fore.RESET)
                        except ValueError:
                            print(Fore.LIGHTRED_EX + '[ERROR] Неверное шк отправления, попробуйте еще раз' + Fore.RESET)
                            continue
                elif task_type == 6:
                    print(Fore.LIGHTWHITE_EX + "[INFO] Вы выбрали [6] - Печать этикетки\n\n" + Fore.WHITE + "(для выхода назад напишите 'Назад')\n\n")
                    while True:
                        try:
                            shipment_id = input(Fore.LIGHTWHITE_EX + "[INPUT] Введите шк отправления(как обычного так и почты россии): ")
                            if str(shipment_id).lower() == 'назад':
                                break
                            shipment_id = int(shipment_id)
                            if len(str(shipment_id)) == 13:
                                shipment_id = parse_code(str(shipment_id))
                            shipment_response = get_shipment(session, shipment_id)
                            if shipment_response.status_code == 404:
                                shipment_response = get_shipment_russian(shipment_id)
                                if shipment_response.json() == {}:
                                    raise ValueError
                                else:
                                    get_sticker_file(session, int(shipment_response.text))
                                    print_sticker_file()
                            else:
                                get_sticker_file(session, int(shipment_id))
                                print_sticker_file()
                        except ValueError:
                            print(Fore.LIGHTRED_EX + '[ERROR] Неверное шк отправления, попробуйте еще раз' + Fore.RESET)
                            continue
                else:
                    raise ValueError
            except ValueError:
                os.system('clear')
                print(Fore.LIGHTRED_EX + 'Неверное число, попробуйте еще раз' + Fore.RESET)
                continue
    except KeyboardInterrupt:
        print(Fore.RED + "\n\n[EXITED] Выполнение программы прекращено")
        exit()


if __name__ == '__main__':
    main()
    # shipments = decode_shipments_from_photo(f'{BASE_DIR}/input/freight_photo.jpg')
    # check_shipments(shipments)
